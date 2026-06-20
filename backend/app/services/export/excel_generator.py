"""生成带格式的 Excel 报表 — 从工具查询的结构化数据生成 .xlsx 文件"""
import os
import re
import uuid
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelGenerator:
    """从工具返回的结构化数据生成带格式的 Excel 报表。

    仅导出 list-of-dict 类型的数据字段 — 标量值（如 completion_rate=85.0）
    已在 AI 的文字回答中体现，不再重复导出。
    """

    # 英文字段名 → 中文列名映射
    FIELD_LABEL_MAP = {
        "user_name": "姓名", "user_code": "学号", "dept_name": "部门",
        "completion_rate": "完成率(%)", "total_courses": "课程总数",
        "courses_completed": "已完成课程", "pass_rate": "通过率(%)",
        "avg_score": "平均分", "total_exams": "考试总数",
        "date": "日期", "value": "数值", "stat_date": "统计日期",
        "metric": "指标", "org_name": "机构名称",
        "class_name": "班级名称", "student_count": "学生人数",
        "course_name": "课程名称", "exam_name": "考试名称",
        "risk_type": "风险类型", "days_since_last_study": "距上次学习天数",
        "avg_composite_score": "综合平均分", "role_name": "角色",
        "count": "数量", "trend": "趋势", "granularity": "粒度",
        "urgent": "是否紧急", "is_at_risk": "是否有风险",
        "error_count": "错误数", "total_error_steps": "总错误步骤",
        "error_rate": "错误率(%)", "rank": "排名",
        "user_id": "用户ID", "dept_code": "部门代码",
        "courseware_code": "课件代码", "exam_code": "考试代码",
        "scope_type": "范围类型", "summary": "摘要",
    }

    # 常见字段名 → Sheet 中文名
    SHEET_NAME_MAP = {
        "breakdown": "明细数据",
        "learners": "学员列表",
        "data_points": "趋势数据",
        "students": "学生列表",
        "classes": "班级列表",
        "top_errors": "错误排行",
        "distribution": "分布统计",
        "org_stats": "机构统计",
    }

    # ── 预定义样式 ──────────────────────────────────────────
    _TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1f1f1f")
    _SUBTITLE_FONT = Font(name="微软雅黑", size=10, color="888888")
    _HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="ffffff")
    _HEADER_FILL = PatternFill(start_color="1f1f1f", end_color="1f1f1f", fill_type="solid")
    _DATA_FONT = Font(name="微软雅黑", size=10)
    _RED_FONT = Font(name="微软雅黑", size=10, color="cc0000")
    _ALT_FILL = PatternFill(start_color="fafafa", end_color="fafafa", fill_type="solid")
    _THIN_BORDER = Border(
        left=Side(style="thin", color="dddddd"),
        right=Side(style="thin", color="dddddd"),
        top=Side(style="thin", color="dddddd"),
        bottom=Side(style="thin", color="dddddd"),
    )

    # ── 公开 API ────────────────────────────────────────────

    def generate(
        self,
        tool_results: list[dict],
        question: str,
        user_scope: str,
        report_dir: str,
    ) -> dict | None:
        """生成格式化的 .xlsx 文件。

        返回元数据 dict，如果没有可导出的表格数据则返回 None。
        """
        sheets_data = self._extract_sheets(tool_results)
        if not sheets_data:
            return None

        os.makedirs(report_dir, exist_ok=True)
        file_id = uuid.uuid4().hex[:12]
        file_name = self._derive_title(question) + ".xlsx"
        safe_name = f"{file_id}.xlsx"
        file_path = os.path.join(report_dir, safe_name)

        wb = Workbook()
        wb.remove(wb.active)  # sheets_data 已保证非空

        for sheet_name, rows in sheets_data:
            ws = wb.create_sheet(title=sheet_name[:31])
            self._write_sheet(ws, sheet_name, rows, file_name, user_scope)

        wb.save(file_path)
        file_size = os.path.getsize(file_path)

        total_rows = sum(len(rows) for _, rows in sheets_data)
        total_columns = max(
            (len(rows[0]) for _, rows in sheets_data if rows), default=0
        )

        return {
            "file_name": file_name,
            "file_url": f"/api/v1/reports/{safe_name}",
            "file_size": file_size,
            "sheets": [
                {
                    "name": name,
                    "rows": len(rows),
                    "columns": len(rows[0]) if rows else 0,
                }
                for name, rows in sheets_data
            ],
            "total_rows": total_rows,
            "total_columns": total_columns,
        }

    # ── 内部方法 ─────────────────────────────────────────────

    def _extract_sheets(
        self, tool_results: list[dict]
    ) -> list[tuple[str, list[dict]]]:
        """从工具结果中提取所有 list-of-dict 字段。

        返回 [(sheet_name, rows), ...] 列表。
        """
        sheets = []
        for tr in tool_results:
            result = tr.get("result", {})
            if not isinstance(result, dict):
                continue
            for key, val in result.items():
                if (
                    isinstance(val, list)
                    and len(val) > 0
                    and isinstance(val[0], dict)
                ):
                    sheet_name = self.SHEET_NAME_MAP.get(
                        key, self._field_to_chinese(key)
                    )
                    sheets.append((sheet_name, val))
        return sheets

    def _field_to_chinese(self, field: str) -> str:
        """英文字段名 → 中文列名，未映射的保留英文"""
        return self.FIELD_LABEL_MAP.get(field, field)

    def _derive_title(self, question: str) -> str:
        """从用户问题中提取报表标题，过滤文件系统非法字符"""
        noise = r"(帮我|请|生成|一个|一份|导出|下载|报表|Excel|excel|数据|的|吗|吧|一下)"
        cleaned = re.sub(noise, "", question).strip()
        if not cleaned:
            cleaned = "数据报表"
        # 过滤文件系统非法字符 (Windows: / \ : * ? " < > |)
        cleaned = re.sub(r'[/\\:*?"<>|]', '_', cleaned)
        return cleaned[:50]

    def _write_sheet(
        self,
        ws,
        sheet_name: str,
        rows: list[dict],
        title: str,
        user_scope: str,
    ):
        """填充单个 Sheet：标题 → 副标题 → 表头 → 数据 → 格式化"""
        if not rows:
            return

        columns = list(rows[0].keys())
        ncols = len(columns)
        nrows = len(rows)

        # 第1行: 标题（合并单元格）
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        title_cell = ws.cell(1, 1, title)
        title_cell.font = self._TITLE_FONT
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # 第2行: 副标题
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        subtitle_cell = ws.cell(2, 1, f"机构: {user_scope} | 生成时间: {now}")
        subtitle_cell.font = self._SUBTITLE_FONT
        subtitle_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 20

        # 第3行: 表头
        for ci, col_name in enumerate(columns, start=1):
            header = self._field_to_chinese(col_name)
            cell = ws.cell(3, ci, header)
            cell.font = self._HEADER_FONT
            cell.fill = self._HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 22
        ws.freeze_panes = "A4"  # 冻结表头

        # 数据行（从第4行开始）
        for ri, row in enumerate(rows, start=4):
            for ci, col_name in enumerate(columns, start=1):
                value = row.get(col_name)
                cell = ws.cell(ri, ci, value)
                cell.font = self._DATA_FONT
                if ri % 2 == 0:  # 偶数行浅灰背景
                    cell.fill = self._ALT_FILL

        # 条件格式: 含"率"字的列 < 60 → 红色
        for ci, col_name in enumerate(columns, start=1):
            header_cn = self._field_to_chinese(col_name)
            if "率" in header_cn:
                for ri in range(4, 4 + nrows):
                    cell = ws.cell(ri, ci)
                    if isinstance(cell.value, (int, float)) and cell.value < 60:
                        cell.font = self._RED_FONT

        # 全表细线边框
        for ri in range(1, 4 + nrows):
            for ci in range(1, ncols + 1):
                ws.cell(ri, ci).border = self._THIN_BORDER

        # 自动列宽（取表头和前20行中最大值）
        for ci, col_name in enumerate(columns, start=1):
            max_len = len(self._field_to_chinese(col_name))
            sample = min(nrows, 20)
            for ri in range(4, 4 + sample):
                val = ws.cell(ri, ci).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 40)
