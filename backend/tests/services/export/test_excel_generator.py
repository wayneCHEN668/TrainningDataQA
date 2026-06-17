"""Tests for ExcelGenerator."""
import os
import tempfile
import pytest
from app.services.export.excel_generator import ExcelGenerator


class TestExtractSheets:
    """测试 _extract_sheets — 从工具结果中过滤 list-of-dict 字段"""

    def test_extracts_list_of_dict_fields(self):
        gen = ExcelGenerator()
        tool_results = [
            {
                "tool_name": "query_completion_rate",
                "result": {
                    "completion_rate": 85.0,
                    "total_learners": 100,
                    "breakdown": [
                        {"dept_name": "信息学院", "rate": 90.0},
                        {"dept_name": "机械学院", "rate": 80.0},
                    ],
                }
            }
        ]
        sheets = gen._extract_sheets(tool_results)
        assert len(sheets) == 1
        name, rows = sheets[0]
        assert "明细" in name  # breakdown → 明细数据
        assert len(rows) == 2
        assert rows[0]["dept_name"] == "信息学院"

    def test_skips_scalar_fields(self):
        gen = ExcelGenerator()
        tool_results = [
            {
                "tool_name": "query_completion_rate",
                "result": {
                    "completion_rate": 85.0,
                    "total_learners": 100,
                    "data_source": "course_grade",
                }
            }
        ]
        sheets = gen._extract_sheets(tool_results)
        assert len(sheets) == 0

    def test_skips_empty_lists(self):
        gen = ExcelGenerator()
        tool_results = [
            {
                "tool_name": "query_incomplete_learners",
                "result": {"incomplete_count": 0, "learners": []},
            }
        ]
        sheets = gen._extract_sheets(tool_results)
        assert len(sheets) == 0

    def test_handles_multiple_tools(self):
        gen = ExcelGenerator()
        tool_results = [
            {
                "tool_name": "query_completion_rate",
                "result": {"breakdown": [{"dept": "A", "val": 1}]},
            },
            {
                "tool_name": "query_learning_trend",
                "result": {"data_points": [{"date": "2026-01-01", "value": 50}]},
            },
        ]
        sheets = gen._extract_sheets(tool_results)
        assert len(sheets) == 2

    def test_handles_multiple_list_fields_in_one_result(self):
        gen = ExcelGenerator()
        tool_results = [
            {
                "tool_name": "query_multi",
                "result": {
                    "breakdown": [{"a": 1}],
                    "learners": [{"b": 2}],
                    "summary": "text only",
                }
            }
        ]
        sheets = gen._extract_sheets(tool_results)
        assert len(sheets) == 2


class TestGenerate:
    """集成测试 — generate() 方法"""

    def test_generates_valid_xlsx(self):
        gen = ExcelGenerator()
        tool_results = [
            {
                "tool_name": "query_completion_rate",
                "result": {
                    "breakdown": [
                        {"dept_name": "信息学院", "completion_rate": 85.5, "total_courses": 12},
                        {"dept_name": "机械学院", "completion_rate": 42.0, "total_courses": 8},
                    ]
                }
            }
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            output = gen.generate(
                tool_results=tool_results,
                question="帮我生成各部门完成率报表",
                user_scope="全部机构",
                report_dir=tmpdir,
            )
            # 检查输出元数据
            assert output is not None
            assert output["file_name"].endswith(".xlsx")
            assert output["file_url"].startswith("/api/v1/reports/")
            assert output["file_size"] > 0
            assert len(output["sheets"]) == 1
            assert output["sheets"][0]["rows"] == 2
            assert output["sheets"][0]["columns"] == 3
            assert output["total_rows"] == 2
            assert output["total_columns"] == 3

            # 检查文件存在于磁盘
            file_name = os.path.basename(output["file_url"])
            file_path = os.path.join(tmpdir, file_name)
            assert os.path.exists(file_path)

            # 验证是有效的 xlsx 文件
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            assert len(wb.sheetnames) == 1
            ws = wb.active
            # 行1: 标题, 行2: 副标题, 行3: 表头, 行4-5: 数据
            assert ws.cell(1, 1).value is not None  # 标题存在
            assert ws.cell(3, 1).value in ("部门", "dept_name")  # 表头行
            assert ws.cell(4, 1).value == "信息学院"
            assert ws.cell(5, 1).value == "机械学院"

    def test_generates_multiple_sheets(self):
        gen = ExcelGenerator()
        tool_results = [
            {
                "tool_name": "query_completion_rate",
                "result": {"breakdown": [{"dept": "A", "rate": 80.0}]},
            },
            {
                "tool_name": "query_learning_trend",
                "result": {"data_points": [{"date": "2026-01-01", "value": 10}]},
            },
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            output = gen.generate(tool_results, "报表", "全部", tmpdir)
            assert output is not None
            assert len(output["sheets"]) == 2

    def test_empty_tool_results_returns_none(self):
        gen = ExcelGenerator()
        with tempfile.TemporaryDirectory() as tmpdir:
            output = gen.generate([], "报表", "全部", tmpdir)
            assert output is None

    def test_conditional_formatting_applied(self):
        """验证完成率 < 60% 的单元格标红"""
        gen = ExcelGenerator()
        tool_results = [
            {
                "tool_name": "query_completion_rate",
                "result": {
                    "breakdown": [
                        {"dept_name": "优秀部门", "completion_rate": 85.0},
                        {"dept_name": "落后部门", "completion_rate": 42.0},
                    ]
                }
            }
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            output = gen.generate(tool_results, "报表", "全部", tmpdir)
            assert output is not None
            from openpyxl import load_workbook
            file_path = os.path.join(tmpdir, os.path.basename(output["file_url"]))
            wb = load_workbook(file_path)
            ws = wb.active
            # 行4列2 = 85.0 (>= 60, 不标红)
            # 行5列2 = 42.0 (< 60, 标红)
            cell_low = ws.cell(5, 2)
            assert cell_low.value == 42.0
            assert cell_low.font.color is not None  # 有颜色设置
